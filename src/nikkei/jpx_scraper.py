"""
jpx_scraper.py
JPX公式サイトから日経225関連のPER/PBRデータを取得するモジュール
https://www.jpx.co.jp/markets/statistics-equities/misc/04.html
"""

import io
import re
from typing import Optional
import requests
import openpyxl
from bs4 import BeautifulSoup

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from src.common.tz import now_jst
from config import PER_LONG_TERM_AVERAGE

JPX_PAGE_URL = "https://www.jpx.co.jp/markets/statistics-equities/misc/04.html"
JPX_BASE_URL = "https://www.jpx.co.jp"

# フォールバック値（スクレイピング失敗時）
FALLBACK_PER = 15.0
FALLBACK_PBR = 1.3


def fetch_per_pbr() -> dict:
    """
    JPX公式サイトから最新のPER/PBRデータを取得する。
    プライム市場総合の加重平均PERを日経225の近似値として使用する。
    EPS は 日経平均終値 ÷ PER で算出する。

    Returns:
        {
            "per": float,          # 加重平均PER（倍）
            "pbr": float,          # 加重平均PBR（倍）
            "per_simple": float,   # 単純平均PER（倍）
            "data_month": str,     # データの対象月（例: "2026年3月"）
            "per_comment": str,    # PER水準コメント
        }
    """
    try:
        # 1. ページからExcelリンクを取得
        xlsx_url = _find_latest_xlsx_url()
        if not xlsx_url:
            raise ValueError("Excelファイルのリンクが見つかりません")

        # 2. Excelをダウンロードして解析
        per_data = _parse_xlsx(xlsx_url)
        if not per_data:
            raise ValueError("Excelからデータを抽出できません")

        # 3. PER水準コメント生成
        per_data["per_comment"] = _generate_per_comment(per_data["per"])

        return per_data

    except Exception as e:
        print(f"[WARN] JPX PER/PBR取得エラー: {e}")
        return _build_fallback()


def calc_eps(nikkei_close: float, per: float) -> float:
    """
    EPS を算出する。EPS = 日経平均終値 ÷ PER
    Args:
        nikkei_close: 日経平均終値
        per: PER（倍）
    Returns:
        EPS（円）
    """
    if per <= 0:
        return 0.0
    return round(nikkei_close / per, 1)


def _find_latest_xlsx_url() -> Optional[str]:
    """JPXページから最新月のExcelファイルURLを取得する"""
    resp = requests.get(JPX_PAGE_URL, timeout=10)
    resp.encoding = resp.apparent_encoding
    soup = BeautifulSoup(resp.text, "html.parser")

    # perpbrYYYYMM.xlsx 形式のリンクを全て取得
    pattern = re.compile(r"perpbr(\d{6})\.xlsx")
    links = []
    for a in soup.find_all("a", href=True):
        match = pattern.search(a["href"])
        if match:
            year_month = match.group(1)
            full_url = JPX_BASE_URL + a["href"] if a["href"].startswith("/") else a["href"]
            links.append((year_month, full_url))

    if not links:
        return None

    # 最新月を選択
    links.sort(key=lambda x: x[0], reverse=True)
    latest = links[0]
    print(f"[INFO] JPX最新データ: {latest[0][:4]}年{latest[0][4:]}月")
    return latest[1]


def _parse_xlsx(url: str) -> Optional[dict]:
    """
    Excelファイルをダウンロードしてプライム市場総合のPER/PBRを抽出する。
    """
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb.active

    # データ行を探索（プライム市場 + 総合/Composite）
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        section = _cell_value(row[1])   # B列: 市場区分名
        category = _cell_value(row[3])  # D列: 種別
        year_month = _cell_value(row[0])  # A列: 年月

        if "プライム" in str(section) and "総合" == str(category).strip():
            per_simple = _parse_number(row[6])  # G列: 単純PER
            pbr_simple = _parse_number(row[7])  # H列: 単純PBR
            per_weighted = _parse_number(row[10])  # K列: 加重PER
            pbr_weighted = _parse_number(row[11])  # L列: 加重PBR

            # 加重PERが取れなければ単純PERを使用
            per = per_weighted if per_weighted else per_simple
            pbr = pbr_weighted if pbr_weighted else pbr_simple

            if per and per > 0:
                # データ月の整形
                data_month = _format_month(year_month)
                print(f"[OK] JPX PER取得: 加重PER={per}倍, 単純PER={per_simple}倍, PBR={pbr}倍")
                return {
                    "per": per,
                    "pbr": pbr or 0.0,
                    "per_simple": per_simple or 0.0,
                    "data_month": data_month,
                }

    return None


def _cell_value(cell):
    """セルの値を取得する（=プレフィクス付きの場合は除去）"""
    val = cell.value
    if val is None:
        return ""
    val = str(val).strip()
    if val.startswith("="):
        val = val[1:]
    return val


def _parse_number(cell) -> Optional[float]:
    """セルから数値を取得する"""
    val = _cell_value(cell)
    if not val or val in ("－", "＊", "-", "*", ""):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _format_month(year_month_str: str) -> str:
    """年月文字列を整形する（例: "2026/03" → "2026年3月"）"""
    try:
        parts = str(year_month_str).split("/")
        if len(parts) == 2:
            year = parts[0]
            month = str(int(parts[1]))
            return f"{year}年{month}月"
    except Exception:
        pass
    return str(year_month_str)


def _generate_per_comment(per: float) -> str:
    """PER水準に基づくコメントを生成する"""
    avg = PER_LONG_TERM_AVERAGE
    if per >= avg + 3:
        return f"PERは長期平均（{avg}倍）を大幅に上回っており、割高水準"
    elif per >= avg + 1:
        return f"PERは長期平均（{avg}倍）をやや上回る"
    elif per >= avg - 1:
        return f"PERは長期平均（{avg}倍）水準で推移"
    elif per >= avg - 3:
        return f"PERは長期平均（{avg}倍）をやや下回る"
    else:
        return f"PERは長期平均（{avg}倍）を大幅に下回っており、割安水準"


def _build_fallback() -> dict:
    """フォールバック値を返す"""
    now = now_jst()
    return {
        "per": FALLBACK_PER,
        "pbr": FALLBACK_PBR,
        "per_simple": FALLBACK_PER,
        "data_month": f"{now.year}年{now.month}月",
        "per_comment": _generate_per_comment(FALLBACK_PER),
    }


if __name__ == "__main__":
    data = fetch_per_pbr()
    print("=== JPX PER/PBR データ ===")
    for k, v in data.items():
        print(f"  {k}: {v}")

    # EPS算出テスト
    nikkei_close = 38000.0
    eps = calc_eps(nikkei_close, data["per"])
    print(f"\n  日経平均={nikkei_close}円, PER={data['per']}倍 → EPS={eps}円")
